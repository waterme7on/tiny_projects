#!/usr/bin/env python
# coding: utf-8

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, colors
fill_red = PatternFill('solid', fgColor='ff0000')


file_path = input('输入文件路径：')
sheet_name = input('输入子表名：')
col_name = input('输入所在列(1,2,...)：')

# file_path = 'trademarks.xlsx'
# sheet_name='test'
# col_name=2

col_name = int(col_name)-1

result_dict = {}

def isRegistered(trademark_id):
    url = 'https://www.trademarkia.com/trademarks-search.aspx?tn='
    html = requests.get(url+trademark_id)
    html_doc = html.text
    soup = BeautifulSoup(html_doc, 'html.parser')
    result = (soup.find_all(attrs={'class':'apply-now row'}))
    # 说明已经被注册了
    return len(result) == 0


data = openpyxl.load_workbook(file_path)
sheet = data[sheet_name]


for i, row in enumerate(sheet.iter_rows(min_row=2)):
    cell = row[col_name]
    isRegi = 0
    if cell.value in result_dict:
        isRegi = result_dict[cell.value]
    else:
        isRegi = isRegistered(str(cell.value))
    if isRegi:
        result_dict[cell.value] = 1
        cell.fill = fill_red
        print(cell.value, '已被注册')
    else:
        result_dict[cell.value] = 0
        print(cell.value, '未被注册')


input('爬取完毕，储存文件：是/否')
data.save(file_path)

