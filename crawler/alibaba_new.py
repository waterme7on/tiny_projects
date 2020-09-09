#!/usr/bin/env python
# coding: utf-8

# In[1]:


# 提供的excel所在路径
file_name='data.xlsx'


# In[ ]:


file_name=input('输入xlsx文件所在路径')
sheet_name=input('输入爬取子表名')


# In[29]:

'''
crawl_columns=input('输入包含的列（链接、选项1、选项2、...、备注对应的列名）')
crawl_columns=crawl_columns.split("\t")
crawl_columns=[c for c in crawl_columns if c != '']
'''
# print(crawl_columns)


# In[ ]:





# In[ ]:





# In[2]:


# 代理的名字和用户密码
# ADSL 账户和密码
# dial_name = "VPN"
# dial_username = "051511259526"
# dial_passwd = "010203"

dial_name = input('输入VPN名')
dial_username = input('输入VPN账号')
dial_passwd = input('输入VPN密码')


# ## 设置代理

# In[3]:



import os
import time
def connect_vpn():
    cmd_str = "rasdial %s %s %s" % (dial_name, dial_username, dial_passwd)
    os.system(cmd_str)
    time.sleep(3)
    
def disconnect_vpn():
    cmd_str = "rasdial %s /disconnect" % dial_name
    os.system(cmd_str)
    time.sleep(1)


# In[4]:



connect_vpn()


# ## 1. 在表中获取各个链接

# In[5]:


import pandas as pd
import numpy as np
import json
import re
from selenium.webdriver.common.proxy import Proxy, ProxyType


# In[6]:


data = pd.read_excel(open(file_name, 'rb'), sheet_name=sheet_name)
data = data.fillna('')


# In[7]:


data.columns


# In[8]:


sub_data = data[['网站链接', '网站选项1', '网站选项2', '网站选项3','网站选项4','网站选项5','网站选项6','网站选项7','网站选项8','网站选项9','网站选项10','网站选项11','网站选项12','网站选项13','网站选项14','网站选项15','网站选项16','备注']]
# sub_data = data[crawl_columns]
sub_data = sub_data.drop_duplicates() # 去重


# In[9]:


sub_data.head()


# In[10]:


# for idx, row in sub_data.iterrows():
#     print(idx)
#     color = row[1]
#     print(color)
#     option = [i for i in row[2:].to_list() if i]
#     print(option)
#     notes = row['备注']
#     print(notes)
#     break
# # 0
# # 1 混彩色猫眼
# # ['4mm', '6mm', '8mm', '10mm', '12mm', '14mm', '16mm', '无']
# # 无


#    

# ## 2. 获得链接后用爬虫爬取

# In[11]:


from selenium import webdriver
from selenium.webdriver.chrome.options import Options


# In[35]:


# connect_vpn()


# In[36]:


# chrome_options = Options()
# # chrome_options.add_argument('--headless')
# # chrome_options.add_argument('--disable-gpu')
# browser = webdriver.Chrome('C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe', options=chrome_options)
# browser.delete_all_cookies()
# browser.maximize_window()
# url='https://detail.1688.com/offer/579617476982.html?spm=a2615.7691456.autotrace-offerGeneral.1.54ca5801IT30Qh'
# browser.get(url)
# browser.implicitly_wait(2)


# In[37]:


# browser.close()


# In[38]:


# while True:
#     try:
#         bar = browser.find_element_by_css_selector('#err > div.tips > p:nth-child(1)')
#         disconnect_vpn()
#         connect_vpn()
#         browser.refresh()
#         browser.implicitly_wait(2)
#     except:
#         # 未遭到检测则停止
#         break


# In[39]:


# # 展开详情页
# try:
#     contents_blk = browser.find_element_by_css_selector("#mod-detail-bd > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content")
#     browser.execute_script("arguments[0].style = 'height: 1000px;';", contents_blk)
#     time.sleep(1)
# except:
#     pass


# In[40]:


#mod-detail-bd > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content > table > tbody > tr:nth-child(1)


# In[41]:


# _options = browser.find_elements_by_css_selector('#mod-detail-bd > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content > table > tbody > tr')
# for _op in _options:
#     if not _op.text:
#         continue
#     _color = (_op.get_attribute('data-sku-config'))
#     _color = json.loads(_color)['skuName']
#     _num = (_op.text.strip('\n-\n+').split(' '))[1]
#     print(_color, _num)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[17]:



# 说明：status_dict记录货品是否有货
# status_dict[货品名]为一个字典
# status_dict[货品名][颜色名] = 1 if 有货 else 0
status_dict = {}

proxy_id = 0

chrome_options = Options()
#chrome_options.add_argument('--headless')
#chrome_options.add_argument('--disable-gpu')
chrome_driverPath = input('输入ChromeDriver所在路径')
# browser = webdriver.Chrome('C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe', options=chrome_options)
browser = webdriver.Chrome(chrome_driverPath, options=chrome_options)
browser.delete_all_cookies()
browser.maximize_window()


# In[21]:



connect_vpn()
# disconnect_vpn()
for idx, row in sub_data.iterrows():
    # 获取该行对应的颜色
    color = row[1]
    status_dict[color] = {}
    # 获取对应颜色下各个型号
    options = [i for i in row[2:-2].to_list() if i]
    notes = row['备注']
    
    # 浏览器爬取
    # 进入网站
    url = row['网站链接']
    if url == '' or not url:
        continue
    if browser.current_url != url:
        browser.get(url)
    # 等待加载
    browser.implicitly_wait(5)

    while True:
        try:
            # 查看是否遭到检测
            # 遭到的换ip
            bar = browser.find_element_by_css_selector('#err')
            print("爬取",color,"遭到检测，换ip")
            disconnect_vpn()
            connect_vpn()
            browser.get(url)
            browser.implicitly_wait(2)
        except:
            # 未遭到检测则停止
            break
    
    print('爬取颜色：',row[1])

    # 只有选项一
    if len(options) == 0:
        # 展开详情页
        try:
            contents_blk = browser.find_element_by_css_selector("#mod-detail-bd > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content")
            browser.execute_script("arguments[0].style = 'height: 1500px;';", contents_blk)
            time.sleep(1)
        except:
            pass
        
        # 获得各个颜色选项
        _options = browser.find_elements_by_css_selector('#mod-detail-bd > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content > table > tbody > tr')
        for _op in _options:
            if not _op.text:
                continue
            _color = (_op.get_attribute('data-sku-config'))
            _color = json.loads(_color)['skuName']
            _num = (_op.text.strip('\n-\n+').split(' '))[1]
            # 不匹配的可以跳过
            if color != _color and color.find(_color) == -1:
                continue
            print(_color, _num)
            
            if int(re.findall(r"\d+\.?\d*",_num)[0]) <= 0:
                # 记录无货信息
                status_dict[color][color] = 0
            else:
                # 记录有货信息
                status_dict[color][color] = int(re.findall(r"\d+\.?\d*",_num)[0])
        print(status_dict[color])
        continue
        
    
    # 有选项一和选项二
    # 获取页面中各个元素
    # 这个获取的部分可能随时会变
    
    # 有图片的
    alters = browser.find_elements_by_css_selector('#mod-detail-bd > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-leading > div.obj-content > ul > li > div > a > span.vertical-img > span > img')
    if not alters:
        alters += browser.find_elements_by_css_selector('#mod-detail-bd > div.detail-v2018-layout-left > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical.dsc-version2018-page-fix-content-mid > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-leading > div.obj-content > ul > li > div > a > span.vertical-img > span > img')
    # 只有文字的
    alters += browser.find_elements_by_css_selector('#mod-detail-bd > div.detail-v2018-layout-left > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical.dsc-version2018-page-fix-content-mid > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-leading > div.obj-content > ul > li > div > a')
    for alter in alters:
        _color = alter.get_attribute('alt')
        if _color == '':
            _color = alter.get_attribute('title')
        if not _color:
            continue
        # 如果当前行的颜色匹配，则点击，否则跳过
        if color != _color and _color.find(color) == -1 and color.find(_color) == -1:
            continue
        print(_color)
        print(options)
        
        try:
            alter.click()
        except:
            continue
        time.sleep(3)

        # 展开详情页
        try:
            contents_blk = browser.find_element_by_css_selector('#mod-detail-bd > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content')
            browser.execute_script("arguments[0].style = 'height: 800px;';", contents_blk)
            time.sleep(1)
        except:
            pass
        # 展开详情页
        try:
            contents_blk = browser.find_element_by_css_selector("#mod-detail-bd > div.detail-v2018-layout-left > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical.dsc-version2018-page-fix-content-mid > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content")
            browser.execute_script("arguments[0].style = 'height: 800px;';", contents_blk)
            time.sleep(1)
        except:
            pass
    
    
        # 得到当前颜色下各个型号
#         _options = browser.find_elements_by_css_selector('#mod-detail-bd > div.detail-v2018-layout-left > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical.dsc-version2018-page-fix-content-mid > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content > table > tbody > tr')
        _options = browser.find_elements_by_css_selector('#mod-detail-bd > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content > table > tbody > tr')
        if not _options:
            _options = browser.find_elements_by_css_selector('#mod-detail-bd > div.detail-v2018-layout-left > div.region-custom.region-detail-property.region-takla.ui-sortable.region-vertical.dsc-version2018-page-fix-content-mid > div.widget-custom.offerdetail_ditto_purchasing > div > div > div > div.obj-sku > div.obj-content > table > tbody > tr')
        # 遍历各个型号
        for _option in _options:
            # 为空则退出
            if not _option.text: 
                continue
            _option = (_option.text.strip('\n-\n+').split(' '))
            print(color, _option)

            # 不在选项中则退出
            # 可能为空块
            # 对size设置过大
            try:
                __op = _option[0]
                __num = _option[len(_option)-1]
            except:
                continue
#             print(color, __op, __num)
            # 如果不在，查看有无间接匹配的
            if __op not in options:
                __oo = ""
                __op_num = int(re.findall(r"\d+\.?\d*",__op)[0])
                for ___op in options:
                    ___op_num = int(re.findall(r"\d+\.?\d*",___op)[0])
                    if ___op_num == __op_num:
                        __oo = ___op
                if not __oo:
                    continue
                __op = __oo

            if int(re.findall(r"\d+\.?\d*",__num)[0]) <= 0:
                # 记录无货信息
                status_dict[color][__op] = 0
            else:
                # 记录有货信息
                status_dict[color][__op] = int(re.findall(r"\d+\.?\d*",__num)[0])

        # 新增：根据备注来进行调节
        # 如果备注否，则说明全部没有货，显示有误
        # 如果备注具体型号，则具体型号没有货
        if notes == '否':
            for option in options:
                status_dict[color][option] = 0
            continue
        elif notes != '':
            empty_options = notes.split(',')
            for option in empty_options:
                option = option.strip(' ')
                status_dict[color][option] = 0
        print(status_dict[color])
        break
disconnect_vpn()


# In[16]:


browser.close()


# In[22]:


# print(status_dict)


# In[23]:


# import json
# with open('data.json', 'w', encoding='utf8') as f:
#     json.dump(status_dict, f)


# ## 取得货品状态后对原excel表格进行更新

# In[24]:


# 对表中颜色进行分析
import openpyxl
from openpyxl.styles import PatternFill, colors

work_book = openpyxl.load_workbook(file_name)
work_sheet = work_book.get_sheet_by_name('产品目录')
print(work_sheet.max_row,work_sheet.max_column)


# In[25]:


fill_red = PatternFill('solid', fgColor='ff0000')
fill_blue = PatternFill('solid', fgColor='0000ff')
fill_yellow = PatternFill('solid', fgColor='ffff00')
fill_white = PatternFill('none', fgColor='ffffff')
# isFirstTime = True

pre_color = {}

# 遍历提供的表格并进行修改
for i in range(1, work_sheet.max_row):
    # 当前行对应的颜色
    row_color = ""
    for j in range(1, work_sheet.max_column):
        _cell = (work_sheet.cell(i, j))
        if not _cell.value:
            continue
        # 如果当前颜色在我们爬取结果中，则记录当前行的颜色
        if _cell.value in status_dict:
            row_color = _cell.value
            if _cell.value not int status_dict[row_color].keys():
                continue
        if row_color == "":
            continue
        # 根据检测之前的颜色和已有状态来变色
        try:
            # 上一次为蓝色的情况
            if _cell.fill == fill_blue:
                # 有货标黄，无货标红
                if _cell.value in status_dict[row_color].keys() and status_dict[row_color][_cell.value]:
                    _cell.fill = fill_yellow
                else:
                    _cell.fill = fill_red
            # 上一次为黄色的情况
            elif _cell.fill == fill_yellow:
                # 有货标白，无货标蓝
                if _cell.value in status_dict[row_color].keys() and status_dict[row_color][_cell.value]:
                    _cell.fill = fill_white
                else:
                    _cell.fill = fill_blue
            # 上一次为红色的情况
            elif _cell.fill == fill_red:
                # 有货标黄，无货标红
                if _cell.value in status_dict[row_color].keys() and status_dict[row_color][_cell.value]:
                    _cell.fill = fill_yellow
                else:
                    _cell.fill = fill_red
            # 上一次为无色的情况
            else:
                if _cell.value in status_dict[row_color].keys() and status_dict[row_color][_cell.value]:
                    # 之前有货，现在有货，依然为无色
                    pass
                else:
                    # 之前有货，现在无货为蓝色
                    _cell.fill = fill_blue
#                     # 如果是第一次，无货则直接置为红色
#                     # 不是第一次无货则置为蓝色
#                     if isFirstTime:
#                         _cell.fill = fill_red
#                     else:
#                         _cell.fill = fill_blue
        except Exception as e:
            pass

work_book.save(file_name)


# In[ ]:




